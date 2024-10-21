# Importações para o Web Bot
from botcity.web import WebBot, Browser, By

# Importações para integração com o SDK do BotCity Maestro
from botcity.maestro import *

# Desativa erros se não estamos conectados ao Maestro
BotMaestroSDK.RAISE_NOT_CONNECTED = False

# Importação para gerenciamento automático do ChromeDriver
from webdriver_manager.chrome import ChromeDriverManager

# Importações para manipulação de datas
from datetime import date, timedelta

# Importações para manipulação de dados
import pandas as pd

# Importações para manipulação de arquivos Excel
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule

# Importação para criação de gráficos
import matplotlib.pyplot as plt

# Importação para parsing de HTML
from bs4 import BeautifulSoup

# Importação para operações com arquivos e diretórios
import os

# Importação para manipulação das teclas do teclado
from selenium.webdriver.common.keys import Keys


def main():
    # Obter informações da execução do Maestro
    maestro = BotMaestroSDK.from_sys_args()
    execution = maestro.get_execution()

    print(f"Task ID is: {execution.task_id}")
    print(f"Task Parameters are: {execution.parameters}")

    # Configurar o WebBot
    bot = WebBot()
    bot.headless = False
    bot.browser = Browser.CHROME
    bot.driver_path = ChromeDriverManager().install()

    try:
        # Coletar taxas do dólar dos últimos 7 dias
        rates = get_dollar_rates(bot)
        
        # Salvar dados em arquivo Excel
        save_excel("cotacao_dolar.xlsx", rates)
        
        # Criar gráfico de tendência
        create_trend_graph("cotacao_dolar.xlsx")
        
        # Salvar o resultado no Maestro
        # maestro.save_result(execution.task_id, f"Dados salvos e gráficos criados com sucesso!")
        
        print("Processo concluído.")
    except Exception as e:
        # Tratar erros e salvar resultado no Maestro
        # maestro.save_result(execution.task_id, f"Erro: {str(e)}")
        print(f"Ocorreu um erro: {str(e)}")
    finally:
        # Encerrar o navegador
        bot.stop_browser()

def scrape_dollar_rate(bot, data_inicio, data_fim):
    try:
        # Verificar se a página foi carregada
        while not bot.find_element('/html/body/app-root/bcb-cookies/div/div/div/div/button[2]', By.XPATH):
            print("A página ainda está carregando.")
            bot.wait(2000)

        # Encontrar o "Prosseguir" para aceitar o cookie da página
        bot.find_element('/html/body/app-root/bcb-cookies/div/div/div/div/button[2]', By.XPATH).click()

        # Encontrar o elemento iframe
        iframe = bot.find_element('/html/body/app-root/app-root/div/div/main/dynamic-comp/div/div[2]/div[1]/div/iframe', By.XPATH)

        # Mudando para o contexto do iframe
        bot.enter_iframe(iframe)

        # Encontrar o campo de data início e inserir a data_inicial
        bot.find_element('#DATAINI', By.CSS_SELECTOR).click()
        for i in range(1, 10, 1):
            bot.type_left()
        for i in range(1, 11, 1):
            bot.delete()

        
        bot.paste(data_inicio.replace('/', ''))

        # Encontrar o campo de data final e inserir a data_final
        bot.find_element('#DATAFIM', By.CSS_SELECTOR).click()
        for i in range(1, 10, 1):
            bot.type_left()
        for i in range(1, 11, 1):
            bot.delete()
        bot.paste(data_fim.replace('/', ''))

        # Encontrar botão de pesquisa e realizar o clique
        bot.find_element('/html/body/div/form/div/input', By.XPATH).click()

        # Verificar se a página foi carregada
        while not bot.find_element('/html/body/div[1]/table', By.XPATH):
            print("A página ainda está carregando.")
            bot.wait(2000)

        # Variável em que será armazenada as datas e a cotação do dolar
        rate_data = []

        # Extrair dados dos resultados
        for i in range(3, 10, 1):
            rate = float(bot.find_element(f'/html/body/div/table/tbody/tr[{i}]/td[3]', By.XPATH).text.replace(',', '.'))
            date = bot.find_element(f'/html/body/div/table/tbody/tr[{i}]/td[1]', By.XPATH).text.strip()

            rate_data.append({
                'date': date,
                'rate': rate
            })
        
        print(rate_data)

        return rate_data

    except Exception as e:
        raise Exception(f"Erro ao obter taxa do dólar: {str(e)}")

def get_dollar_rates(bot):
    url = "https://www.bcb.gov.br/estabilidadefinanceira/historicocotacoes"
    rates = {}

    try:
        # Navegar até a página de cotações do Banco Central
        bot.browse(url)

        # Coletar taxas dos últimos 7 dias

        # Data atual
        data_fim = date.today()
        print(f'Data início: {data_fim}')

        # Data final (subtraindo 6 dias)
        data_inicio = data_fim - timedelta(days=8)
        print(f'Data final: {data_inicio}')

        # Formatar as datas no formato dd/mm/aaaa
        data_fim_str = data_fim.strftime("%d/%m/%Y")
        data_inicio_str = data_inicio.strftime("%d/%m/%Y")

        print(f'Data início formatada: {data_fim_str}')
        print(f'Data final formatada: {data_inicio_str}')
        
        rates = scrape_dollar_rate(bot, data_inicio_str, data_fim_str)
    
    except Exception as e:
        print(f"Erro ao coletar taxas: {str(e)}")
        raise  # Relançar a exceção para que o bloco finally do main funcione corretamente

    return rates

def save_excel(file_path, rates):
    try:
        # Abrir ou criar um novo arquivo Excel
        if os.path.exists(file_path):
            wb = load_workbook(filename=file_path)
        else:
            wb = Workbook()

        sheet = wb.active

        # Inserir cabeçalhos
        sheet.cell(row=1, column=1).value = "Data"
        sheet.cell(row=1, column=2).value = "Taxa do Dólar"

        # Inserir dados
        for i, rate_data in enumerate(rates, start=2):
            sheet.cell(row=i, column=1).value = rate_data['date']
            sheet.cell(row=i, column=2).value = rate_data['rate']

        # Criar gráfico de barras
        chart = BarChart()
        data = Reference(sheet, min_col=2, min_row=1, max_row=len(rates)+1)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=len(rates)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.title = 'Variação Diária do Dólar'

        # Configurar eixos
        chart.x_axis.title = 'Data'
        chart.y_axis.title = 'Taxa do Dólar'

        # Adicionar o gráfico à folha
        sheet.add_chart(chart, 'E2')

        # Adicionar regras de formatação condicional
        color_scale = ColorScaleRule(start_type='min', start_color='FF99CC00', end_type='max', end_color='FFFF0000')
        sheet.conditional_formatting.add(f'B2:B{len(rates)+1}', color_scale)

        # Salvar o arquivo Excel
        wb.save(file_path)

    except Exception as e:
        raise Exception(f"Erro ao salvar o Excel: {str(e)}")

def create_trend_graph(file_path):
    try:
        # Ler os dados do arquivo Excel
        df = pd.read_excel(file_path)

        # Corrigir a conversão das datas para o formato adequado
        df['Data'] = pd.to_datetime(df['Data'], format='%d/%m/%Y')  # Formato corrigido
        rates = df['Taxa do Dólar'].astype(float)
        
        # Criar um novo gráfico
        plt.figure(figsize=(12, 6))
        plt.plot(df['Data'], rates, marker='o')  # Adicionando marcadores para melhor visualização
        plt.title('Tendência do Dólar ao Longo do Tempo')
        plt.xlabel('Data')
        plt.ylabel('Taxa do Dólar')
        plt.grid(True)
        plt.tight_layout()
        
        # Salvar o gráfico como imagem
        graph_file = os.path.splitext(file_path)[0] + '_trend.png'
        plt.savefig(graph_file)
        plt.close()  # Fechar a figura para liberar memória

    except Exception as e:
        raise Exception(f"Erro ao criar gráfico de tendência: {str(e)}")

if __name__ == "__main__":
    main()
