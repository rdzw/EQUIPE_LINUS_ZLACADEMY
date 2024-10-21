# Importações para o Web Bot
from botcity.web import WebBot, Browser, By

# Importações para integração com o SDK do BotCity Maestro
from botcity.maestro import *

# Desativa erros se não estamos conectados ao Maestro
BotMaestroSDK.RAISE_NOT_CONNECTED = False

# Importação para gerenciamento automático do ChromeDriver
from webdriver_manager.chrome import ChromeDriverManager

# Importações para requisições HTTP
import requests

# Importações para manipulação de datas
from datetime import date, timedelta

# Importações para manipulação de dados
import pandas as pd

# Importações para manipulação de arquivos Excel
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule

# Importação para criação de gráficos
import matplotlib.pyplot as plt

# Importação para parsing de HTML
from bs4 import BeautifulSoup

# Importação para operações com arquivos e diretórios
import os

def main():
    # Obter informações da execução do Maestro
    maestro = BotMaestroSDK.from_sys_args()
    execution = maestro.get_execution()

    print(f"Task ID is: {execution.task_id}")
    print(f"Task Parameters are: {execution.parameters}")

    # Configurar o WebBot
    bot = WebBot()
    bot.headless = False
    bot.browser = Browser.CHROME
    bot.driver_path = ChromeDriverManager().install()

    try:
        # Coletar taxas do dólar dos últimos 7 dias
        rates = get_dollar_rates(bot)
        
        # Salvar dados em arquivo Excel
        save_excel("cotacao_dolar.xlsx", rates)
        
        # Criar gráfico de tendência
        create_trend_graph("cotacao_dolar.xlsx")
        
        # Salvar o resultado no Maestro
        # maestro.save_result(execution.task_id, f"Dados salvos e gráficos criados com sucesso!")
        
        print("Processo concluído.")
    except Exception as e:
        # Tratar erros e salvar resultado no Maestro
        # maestro.save_result(execution.task_id, f"Erro: {str(e)}")
        print(f"Ocorreu um erro: {str(e)}")
    finally:
        # Encerrar o navegador
        bot.stop_browser()

def scrape_dollar_rate(bot, data_inicial, data_final):
    try:
        while bot.find_element('//*[@id="DATAINI"]', By.XPATH) is None:
            bot.wait(1000)
            print('Carregando...')
        
        # Encontrar o campo de data início
        start_date_input = bot.find_element('//*[@id="DATAINI"]', By.XPATH)

        # Encontrar o campo de data final
        final_date_input = bot.find_element('//*[@id="DATAFIM"]', By.XPATH)

        # Preencher campos de datas
        start_date_input.send_keys(data_inicial)
        final_date_input.send_keys(data_final)

        # Encontrar botão de pesquisa
        search_button = bot.find_element('/html/body/div/form/div/input', By.XPATH)
        search_button.click()

        while bot.find_element(f'/html/body/div/table/tbody/tr[{i}]/td[3]', By.XPATH) is None:
            bot.wait(1000)
            print('Carregando...')

        # Variável em que será armazenada as datas e a cotação do dolar
        rate_data = []

        # Extrair dados dos resultados
        for i in range(3, 9, 1):
            rate = float(bot.find_element(f'/html/body/div/table/tbody/tr[{i}]/td[3]', By.XPATH).text.replace(',', '.'))
            date = bot.find_element(f'/html/body/div/table/tbody/tr[{i}]/td[1]', By.XPATH).text.strip()

            rate_data.append({
                'date': date,
                'rate': rate
            })

        return rate_data

    except Exception as e:
        raise Exception(f"Erro ao obter taxa do dólar: {str(e)}")

def get_dollar_rates(bot):
    url = "https://www.bcb.gov.br/estabilidadefinanceira/historicocotacoes"
    rates = {}

    # Navegar até a página de cotações do Banco Central
    bot.browse(url)

    # Data atual
    data_inicial = date.today()

    # Data final (subtraindo 6 dias)
    data_final = data_inicial - timedelta(days=6)

    # Formatar as datas no formato dd/mm/aaaa
    data_inicial_str = data_inicial.strftime("%d/%m/%Y")
    data_final_str = data_final.strftime("%d/%m/%Y")
    
    rates = scrape_dollar_rate(bot, data_inicial_str, data_final_str)
        
    return rates

def save_excel(file_path, rates):
    # Abrir ou criar um novo arquivo Excel
    if os.path.exists(file_path):
        wb = load_workbook(filename=file_path)
    else:
        wb = Workbook()
    
    sheet = wb.active
    
    # Inserir cabeçalhos
    sheet.cell(row=1, column=1).value = "Data"
    sheet.cell(row=1, column=2).value = "Taxa do Dólar"
    
    # Inserir dados
    for i, rate_data in enumerate(rates, start=2):
        sheet.cell(row=i, column=1).value = rate_data['date']
        sheet.cell(row=i, column=2).value = rate_data['rate']

    
    # Criar gráfico de barras
    chart = BarChart()
    data = Reference(sheet, min_col=2, min_row=1, max_row=len(rates)+1)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=len(rates)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = 'Variação Diária do Dólar'
    
    # Configurar eixos
    chart.x_axis.title = 'Data'
    chart.y_axis.title = 'Taxa do Dólar'
    
    # Adicionar o gráfico à folha
    sheet.add_chart(chart, 'E2')
    
    # Adicionar regras de formatação condicional
    color_scale = ColorScaleRule(start_type='min', start_color='FF99CC00', end_type='max', end_color='FFFF0000')
    sheet.conditional_formatting.add(f'B2:B{len(rates)+1}', color_scale)
    
    # Salvar o arquivo Excel
    wb.save(file_path)

def create_trend_graph(file_path):
    # Ler os dados do arquivo Excel
    df = pd.read_excel(file_path)
    dates = pd.to_datetime(df['Data'], format='%Y-%m-%d')
    rates = df['Taxa do Dólar'].astype(float)
    
    # Criar um novo gráfico
    plt.figure(figsize=(12, 6))
    plt.plot(dates, rates)
    plt.title('Tendência do Dólar ao Longo do Tempo')
    plt.xlabel('Data')
    plt.ylabel('Taxa do Dólar')
    plt.grid(True)
    plt.tight_layout()
    
    # Salvar o gráfico como imagem
    graph_file = os.path.splitext(file_path)[0] + '_trend.png'
    plt.savefig(graph_file)
    plt.close()

if __name__ == '__main__':
    main()
